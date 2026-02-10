import os
import sys
import traceback
import argparse
from pathlib import Path
import pandas as pd
from bca import BCAParser
from mandiri import MandiriParser
from cimb import CIMBParser

def process_file(file_path: Path, output_dir: Path, password: str = None):
    """
    Process a single bank statement file and export to CSV.
    
    Args:
        file_path: Path to the bank statement file
        output_dir: Directory to save the output CSV
        password: Optional password for protected Excel files
    """
    filename = file_path.name.lower()
    parser = None
    account_owner = "Unknown"
    
    print(f"Inspecting {filename}...")

    # Enhanced parser selection using same logic as file finding
    if filename.endswith('.pdf'):
        # Check for BCA (account number pattern or 'bca' keyword or month patterns)
        is_bca = 'bca' in filename or any(month in filename for month in ['_jul_', '_agust_', '_sept_', '_okt_', '_nov_', '_des_', '_jan_', '_feb_', '_mar_', '_apr_', '_mei_', '_jun_'])
        # Check for CIMB (CASA or cimb keyword)
        is_cimb = 'cimb' in filename or 'casa' in filename
        
        if is_bca and not is_cimb:
            parser = BCAParser(str(file_path), account_owner)
            print(f"Selected BCAParser for {filename}")
        elif is_cimb:
            parser = CIMBParser(str(file_path), account_owner)
            print(f"Selected CIMBParser for {filename}")
    elif filename.endswith('.xlsx') or filename.endswith('.xls'):
        # Check for Mandiri (e-statement or mandiri keyword)
        if 'e-statement' in filename or 'mandiri' in filename:
            parser = MandiriParser(str(file_path), account_owner, password=password)
            print(f"Selected MandiriParser for {filename}")
    
    if not parser:
        print(f"No parser found for {filename}")
        return
    if parser:
        print(f"Processing {filename}...")
        try:
            transactions = parser.parse()
            if not transactions:
                print(f"No transactions found in {filename}")
                return

            # Determine date range
            dates = [t.date for t in transactions if t.date]
            if not dates:
                print(f"No valid dates found in {filename}")
                return

            min_date = min(dates)
            max_date = max(dates)
            
            # Format: Bank-[Owner Name]-AccountNumber-Dates.csv
            acct_num = parser.account_number if parser.account_number != "Unknown" else "unknown"
            
            # Use extracted name, or fallback to "Account Owner" if extraction failed
            owner_name = parser.account_owner if parser.account_owner != "Unknown" else "Account Owner"
            
            # Sanitize owner name for filename (remove illegal chars)
            owner_safe = "".join([c for c in owner_name if c.isalpha() or c.isspace() or c in ['.']]).strip()
            
            bank_name = ""
            if isinstance(parser, BCAParser):
                bank_name = "BCA"
            elif isinstance(parser, MandiriParser):
                bank_name = "Mandiri"
            elif isinstance(parser, CIMBParser):
                bank_name = "CIMB"
            else:
                bank_name = "Bank"
             
            start_str = min_date.strftime("%d%m%Y")
            end_str = max_date.strftime("%d%m%Y")
            
            output_filename = f"{bank_name}-[{owner_safe}]-{acct_num}-{start_str}-{end_str}.csv"
            output_path = output_dir / output_filename
            
            data = [t.to_dict() for t in transactions]
            df = pd.DataFrame(data)
            
            cols = ["Date", "Description", "Reference No", "Debit", "Credit", "Balance", "Bank", "Owner"]
            for c in cols:
                if c not in df.columns:
                    df[c] = None
            df = df[cols]
            
            df.to_csv(output_path, index=False)
            print(f"Exported {len(df)} transactions to {output_path}")
            
        except Exception as e:
            print(f"Error processing {filename}:")
            traceback.print_exc()

def find_bank_files(folder_path: Path, bank_name: str, password: str = None) -> list:
    """
    Recursively find all bank statement files matching the specified bank.
    
    Detects both default download names and custom-named files:
    - BCA: Account number pattern (e.g., [ACCOUNT_NUMBER]_JUL_2025.pdf)
    - Mandiri: e-Statement prefix (e.g., e-Statement_[ACCOUNT_NUMBER]_dates.xlsx)
    - CIMB: CASA or CIMB keywords (e.g., CASA_Statement_Nov2025.pdf)
    
    Uses multiple validation criteria to reduce false positives:
    - Filename pattern match
    - File extension validation
    - File content check for bank-specific keywords
    
    Args:
        folder_path: Root folder to search
        bank_name: Bank name ('bca', 'mandiri', or 'cimb')
        password: Optional password for Mandiri files
    
    Returns:
        List of file paths matching the bank criteria
    """
    bank_name = bank_name.lower()
    matching_files = []
    
    # Define validation criteria for each bank
    # Includes both default download names and custom naming patterns
    patterns = {
        'bca': {
            'extensions': ('.pdf',),
            'filename_patterns': ['bca', '_jul_', '_agust_', '_sept_', '_okt_', '_nov_', '_des_', '_jan_', '_feb_', '_mar_', '_apr_', '_mei_', '_jun_', '_maret_', '_oktober_', '_desember_'],  # BCA uses account_MONTH_YEAR pattern
            'keywords': ['bca', 'bank central', 'mutasi rekening', 'saldo', 'rekening'],
        },
        'mandiri': {
            'extensions': ('.xlsx', '.xls'),
            'filename_patterns': ['mandiri', 'e-statement'],  # Mandiri default is e-Statement
            'keywords': ['mandiri', 'bank mandiri', 'laporan transaksi', 'tanggal', 'uraian'],
        },
        'cimb': {
            'extensions': ('.pdf',),
            'filename_patterns': ['cimb', 'casa'],  # CIMB uses CASA for their account type
            'keywords': ['cimb', 'casa', 'bank cimb', 'mutasi rekening'],
        }
    }
    
    if bank_name not in patterns:
        print(f"Error: Unknown bank '{bank_name}'. Supported banks: bca, mandiri, cimb")
        return []
    
    config = patterns[bank_name]
    extensions = config['extensions']
    filename_patterns = config['filename_patterns']
    keywords = config['keywords']
    
    # Search recursively
    print(f"Searching for {bank_name.upper()} files in {folder_path}...")
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            file_lower = file.lower()
            file_path = Path(root) / file
            
            # Step 1: Check filename contains any bank pattern
            matches_pattern = any(pattern in file_lower for pattern in filename_patterns)
            if not matches_pattern:
                continue
            
            # Step 2: Check file extension
            if not any(file_lower.endswith(ext) for ext in extensions):
                continue
            
            # Step 3: Basic file content validation
            is_valid = False
            try:
                if file_lower.endswith('.pdf'):
                    # Quick check: read first page of PDF
                    try:
                        import pdfplumber
                        with pdfplumber.open(file_path) as pdf:
                            first_page = pdf.pages[0].extract_text() if pdf.pages else ""
                            content_lower = first_page.lower() if first_page else ""
                            # Check if any keyword is found in first page
                            is_valid = any(kw in content_lower for kw in keywords)
                    except Exception as e:
                        print(f"  Warning: Could not read PDF {file_path} - {type(e).__name__}")
                        continue
                
                elif file_lower.endswith(('.xlsx', '.xls')):
                    # Quick check: read first sheet
                    try:
                        from openpyxl import load_workbook
                        wb = load_workbook(file_path, data_only=True)
                        ws = wb.active
                        # Check content from first 20 rows
                        content = ""
                        for row in list(ws.iter_rows(values_only=True))[:20]:
                            content += " ".join(str(cell) if cell else "" for cell in row).lower() + " "
                        is_valid = any(kw in content for kw in keywords)
                        wb.close()
                    except Exception as e:
                        # If file fails to open (likely password-protected), trust filename pattern matching
                        # The actual parser will handle the password
                        is_valid = True  # Accept based on filename pattern matching
                
                if is_valid:
                    matching_files.append(file_path)
                    print(f"  ✓ Valid statement: {file_path}")
                else:
                    print(f"  ✗ Skipped (not a valid {bank_name.upper()} statement): {file_path}")
            
            except Exception as e:
                print(f"  ✗ Skipped (error validating): {file_path} - {str(e)}")
                continue
    
    return matching_files

def main():
    parser = argparse.ArgumentParser(
        description="Process bank statement files and export to CSV format.",
        epilog="Example: python process_statements.py /path/to/statements bca"
    )
    
    parser.add_argument(
        'folder',
        help='Path to folder containing bank statement files'
    )
    
    parser.add_argument(
        'bank',
        nargs='?',
        default='all',
        help='Bank name: bca, mandiri, cimb, or all (default: all)'
    )
    
    parser.add_argument(
        '-o', '--output',
        help='Output directory (default: ./output)',
        default=None
    )
    
    parser.add_argument(
        '-p', '--password',
        help='Password for protected Excel files (Mandiri)',
        default=None
    )
    
    args = parser.parse_args()
    
    # Validate folder path
    source_dir = Path(args.folder)
    if not source_dir.exists():
        print(f"Error: Folder not found: {source_dir}")
        sys.exit(1)
    
    if not source_dir.is_dir():
        print(f"Error: Path is not a directory: {source_dir}")
        sys.exit(1)
    
    # Set output directory
    if args.output:
        output_dir = Path(args.output)
    else:
        output_dir = Path(__file__).parent / "output"
    
    output_dir.mkdir(exist_ok=True, parents=True)
    print(f"Output directory: {output_dir}\n")
    
    # Find files
    bank_arg = args.bank.lower()
    
    if bank_arg == 'all':
        banks = ['bca', 'mandiri', 'cimb']
    else:
        banks = [bank_arg]
    
    total_processed = 0
    for bank in banks:
        # Pass password for Mandiri files
        files = find_bank_files(source_dir, bank, password=args.password if bank == 'mandiri' else None)
        
        if not files:
            print(f"No {bank.upper()} files found.\n")
            continue
        
        print(f"Processing {len(files)} {bank.upper()} file(s)...\n")
        
        for file_path in files:
            process_file(file_path, output_dir, password=args.password if bank == 'mandiri' else None)
            total_processed += 1
            print()
    
    if total_processed == 0:
        print("No matching bank statement files were found.")
        sys.exit(1)
    else:
        print(f"✓ Successfully processed {total_processed} file(s)")

if __name__ == "__main__":
    main()
